#!/usr/bin/env python
#
# Generate report in Excel format (from xml input)
# 

import sys,os,shelve
import re,dfxml,fiwalk
from bc_utils import filename_from_path

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

from openpyxl.cell import get_column_letter

def build_local_wb(ws, fi, row_idx):
    ws.cell('%s%s'%('A', row_idx)).value = '%s' % fi.filename()
    ws.cell('%s%s'%('B', row_idx)).value = '%s' % fi.ext()
    ws.cell('%s%s'%('C', row_idx)).value = '%s' % str(fi.filesize())
    ws.cell('%s%s'%('D', row_idx)).value = '%s' % str(fi.mode())
    ws.cell('%s%s'%('E', row_idx)).value = '%s' % str(fi.ctime())
    ws.cell('%s%s'%('F', row_idx)).value = '%s' % str(fi.atime())
    ws.cell('%s%s'%('G', row_idx)).value = '%s' % str(fi.crtime())
    ws.cell('%s%s'%('H', row_idx)).value = '%s' % str(fi.mtime())
    ws.cell('%s%s'%('I', row_idx)).value = '%s' % str(fi.md5())
    ws.cell('%s%s'%('J', row_idx)).value = '%s' % str(fi.sha1())

def process_files(fn, ws):

    row_idx = [2]

    # Callback function to process the SAX stream
    def cb(fi):
        # add the md5 to the set
        if fi.is_file() and fi.filesize():
            ext = fi.ext()
            if ext:
                build_local_wb(ws, fi, row_idx[0])
                row_idx[0] += 1

    if fn.endswith('xml'):
        # We use this call if we're processing a fiwalk XML file
        fiwalk.fiwalk_using_sax(xmlfile=open(fn, 'rb'),callback=cb)
    else:
        # We use this call if we're processing a disk image
        fiwalk.fiwalk_using_sax(imagefile=open(fn, 'rb'),callback=cb)

def bc_generate_xlsx(fn):

    wb = Workbook()
    #dest_filename = r'test_book.xlsx'
    dest_filename = filename_from_path(fn) + ".xlsx"
    print("Generating Excel report ", dest_filename)
    ws = wb.worksheets[0]
    ws.title = "File Object Information"

    ws.cell('%s%s'%('A', '1')).value = '%s' % "Filename"
    ws.cell('%s%s'%('B', '1')).value = '%s' % "Extension"
    ws.cell('%s%s'%('C', '1')).value = '%s' % "Filesize"
    ws.cell('%s%s'%('D', '1')).value = '%s' % "Mode"
    ws.cell('%s%s'%('E', '1')).value = '%s' % "Change time"
    ws.cell('%s%s'%('F', '1')).value = '%s' % "Access time"
    ws.cell('%s%s'%('G', '1')).value = '%s' % "Create time"
    ws.cell('%s%s'%('H', '1')).value = '%s' % "Modification time"
    ws.cell('%s%s'%('I', '1')).value = '%s' % "MD5 Hash" 
    ws.cell('%s%s'%('J', '1')).value = '%s' % "SHA1 Hash"

    process_files(fn, ws)

    wb.save(filename=dest_filename)

